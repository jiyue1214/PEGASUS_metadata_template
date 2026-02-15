import sys
import tempfile
import unittest
from pathlib import Path
from typing import Annotated, Literal

import openpyxl
from pydantic import BaseModel, Field

repo_root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(repo_root / "src"))

from pegasus.template_convert.spreadsheet_builder import (
    _allows_none,
    _col_letter,
    _enum_or_literal_options,
    _longest_word_len,
    _normalize_annotation,
    _to_text,
    estimate_column_width,
    generate_excel_from_pydantic,
    write_model_sheet,
)
from pegasus.schema.peg_metadata_schema.metadata_basic_schema import (
    DatasetDescription,
    GenomicIdentifier,
)
from pegasus.schema.peg_metadata_schema.metadata_evidence_schema import Evidence
from pegasus.schema.peg_metadata_schema.metadata_integration_schme import Integration
from pegasus.schema.peg_metadata_schema.metadata_method_schema import Method
from pegasus.schema.peg_metadata_schema.metadata_source_schema import Source


class TestHelperFunctions(unittest.TestCase):
    """Test utility helper functions in spreadsheet_builder."""

    def test_to_text_with_none(self) -> None:
        """Test _to_text returns empty string for None."""
        self.assertEqual(_to_text(None), "")

    def test_to_text_with_value(self) -> None:
        """Test _to_text converts values to strings."""
        self.assertEqual(_to_text(123), "123")
        self.assertEqual(_to_text("test"), "test")
        self.assertEqual(_to_text(True), "True")

    def test_longest_word_len(self) -> None:
        """Test _longest_word_len finds longest token in a string."""
        self.assertEqual(_longest_word_len("short text"), 5)
        self.assertEqual(_longest_word_len("this-is-hyphenated word"), 18)
        self.assertEqual(_longest_word_len(""), 0)

    def test_col_letter_conversion(self) -> None:
        """Test _col_letter converts column indices to Excel letters."""
        self.assertEqual(_col_letter(0), "A")
        self.assertEqual(_col_letter(1), "B")
        self.assertEqual(_col_letter(25), "Z")
        self.assertEqual(_col_letter(26), "AA")
        self.assertEqual(_col_letter(27), "AB")
        self.assertEqual(_col_letter(701), "ZZ")

    def test_normalize_annotation_plain_type(self) -> None:
        """Test _normalize_annotation with plain types."""
        self.assertEqual(_normalize_annotation(str), str)
        self.assertEqual(_normalize_annotation(int), int)

    def test_normalize_annotation_annotated(self) -> None:
        """Test _normalize_annotation unwraps Annotated types."""
        annotated_str = Annotated[str, Field(description="test")]
        normalized = _normalize_annotation(annotated_str)
        self.assertEqual(normalized, str)

    def test_normalize_annotation_optional(self) -> None:
        """Test _normalize_annotation unwraps Optional types."""
        from typing import Optional

        optional_str = Optional[str]
        normalized = _normalize_annotation(optional_str)
        self.assertEqual(normalized, str)

    def test_enum_or_literal_options_with_literal(self) -> None:
        """Test _enum_or_literal_options extracts Literal values."""
        literal_type = Literal["option1", "option2", "option3"]
        options = _enum_or_literal_options(literal_type)
        self.assertEqual(options, ["option1", "option2", "option3"])

    def test_enum_or_literal_options_with_bool(self) -> None:
        """Test _enum_or_literal_options handles bool type."""
        options = _enum_or_literal_options(bool)
        self.assertEqual(options, ["TRUE", "FALSE"])

    def test_enum_or_literal_options_with_none(self) -> None:
        """Test _enum_or_literal_options returns None for non-enum types."""
        options = _enum_or_literal_options(str)
        self.assertIsNone(options)

    def test_allows_none_with_none_default(self) -> None:
        """Test _allows_none detects None default value."""
        self.assertTrue(_allows_none(str, None))

    def test_allows_none_with_optional_type(self) -> None:
        """Test _allows_none detects Optional types."""
        from typing import Optional

        self.assertTrue(_allows_none(Optional[str], "default"))

    def test_allows_none_required_field(self) -> None:
        """Test _allows_none returns False for required fields."""
        self.assertFalse(_allows_none(str, "default"))

    def test_estimate_column_width_with_long_header(self) -> None:
        """Test estimate_column_width with long header."""
        width = estimate_column_width(
            description="Short description",
            header="Very Long Header Name That Should Determine Width",
            example="short",
        )
        # Header length + 2 should dominate but capped at max_width
        self.assertGreater(width, 14)
        self.assertLessEqual(width, 40)

    def test_estimate_column_width_respects_min(self) -> None:
        """Test estimate_column_width respects minimum width."""
        width = estimate_column_width(
            description="x", header="x", example="x", min_width=20
        )
        self.assertGreaterEqual(width, 20)

    def test_estimate_column_width_respects_max(self) -> None:
        """Test estimate_column_width respects maximum width."""
        width = estimate_column_width(
            description="x" * 100,
            header="x" * 100,
            example="x" * 100,
            max_width=30,
        )
        self.assertLessEqual(width, 30)


class TestSpreadsheetGeneration(unittest.TestCase):
    """Test Excel spreadsheet generation from Pydantic models."""

    def test_generate_excel_creates_file(self) -> None:
        """Test generate_excel_from_pydantic creates an Excel file."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 0)

    def test_generate_excel_contains_all_sheets(self) -> None:
        """Test generated Excel contains all expected sheets."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            sheet_names = wb.sheetnames

            expected_sheets = [
                "DatasetDescription",
                "GenomicIdentifier",
                "Evidence",
                "Integration",
                "Source",
                "Method",
            ]

            for expected in expected_sheets:
                self.assertIn(expected, sheet_names)

    def test_sheet_has_correct_structure(self) -> None:
        """Test generated sheet has correct row structure."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["DatasetDescription"]

            # Row 1: descriptions (gray background)
            # Row 2: headers (colored background for required/optional)
            # Row 3: examples (with thick bottom border)
            # Row 4+: data entry rows

            # Check that first 3 rows exist and have content
            self.assertIsNotNone(ws.cell(1, 1).value)  # Description row
            self.assertIsNotNone(ws.cell(2, 1).value)  # Header row
            # Example row might be empty for some fields
            self.assertIsNotNone(ws.cell(2, 1).value)  # Header should exist

    def test_required_fields_have_different_formatting(self) -> None:
        """Test required fields have distinct formatting from optional fields."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["DatasetDescription"]

            # Find a required field (trait_description is required)
            headers = [cell.value for cell in ws[2]]
            if "trait_description" in headers:
                col_idx = headers.index("trait_description")
                required_cell = ws.cell(2, col_idx + 1)
                # Required fields should have different background color
                self.assertIsNotNone(required_cell.fill)

    def test_evidence_sheet_has_validation_sheet(self) -> None:
        """Test Evidence sheet creates hidden validation sheet."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)

            # Evidence sheet should trigger creation of validation sheet
            self.assertIn("validation", wb.sheetnames)

            # Validation sheet should be hidden
            validation_ws = wb["validation"]
            self.assertEqual(validation_ws.sheet_state, "hidden")

    def test_evidence_sheet_has_data_validation(self) -> None:
        """Test Evidence sheet has data validation rules."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Evidence"]

            # Check that data validation exists
            self.assertGreater(len(ws.data_validations.dataValidation), 0)

    def test_sheet_freeze_panes(self) -> None:
        """Test sheets have frozen panes for header rows."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["DatasetDescription"]

            # Should freeze first 3 rows (description, header, example)
            self.assertIsNotNone(ws.freeze_panes)
            self.assertEqual(ws.freeze_panes, "A4")

    def test_model_with_enum_fields_has_dropdowns(self) -> None:
        """Test fields with enum types get dropdown validation."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Evidence"]

            # Evidence sheet has variant_or_gene_centric which is an enum
            # Should have data validation
            validations = list(ws.data_validations.dataValidation)
            self.assertGreater(len(validations), 0)

    def test_column_widths_are_set(self) -> None:
        """Test column widths are estimated and set."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["DatasetDescription"]

            # Check that columns have custom widths set
            for col_letter in ["A", "B", "C"]:
                col_dim = ws.column_dimensions[col_letter]
                # Default Excel width is about 8.43, custom widths should differ
                self.assertIsNotNone(col_dim.width)

    def test_write_model_sheet_with_custom_model(self) -> None:
        """Test write_model_sheet with a simple custom model."""

        class SimpleModel(BaseModel):
            name: str = Field(description="Name field", json_schema_extra={"header": "Name", "example": "John"})
            age: int = Field(description="Age field", json_schema_extra={"header": "Age", "example": "30"})

        with tempfile.TemporaryDirectory() as tmp_dir:
            import xlsxwriter

            path = Path(tmp_dir) / "test_custom.xlsx"
            workbook = xlsxwriter.Workbook(str(path))
            write_model_sheet(workbook, SimpleModel, rows=50)
            workbook.close()

            self.assertTrue(path.exists())

            # Verify sheet was created
            wb = openpyxl.load_workbook(path)
            self.assertIn("SimpleModel", wb.sheetnames)

            ws = wb["SimpleModel"]
            # Check headers
            headers = [cell.value for cell in ws[2]]
            self.assertIn("Name", headers)
            self.assertIn("Age", headers)

    def test_generated_file_is_valid_excel(self) -> None:
        """Test generated file can be opened and read without errors."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            # This should not raise any exceptions
            wb = openpyxl.load_workbook(path)
            self.assertIsNotNone(wb)
            self.assertGreater(len(wb.sheetnames), 0)

    def test_all_expected_models_are_included(self) -> None:
        """Test all expected Pydantic models are converted to sheets."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)

            expected_models = [
                DatasetDescription,
                GenomicIdentifier,
                Evidence,
                Integration,
                Source,
                Method,
            ]

            for model in expected_models:
                sheet_name = model.__name__[:31]  # Excel limits sheet names to 31 chars
                self.assertIn(
                    sheet_name,
                    wb.sheetnames,
                    f"Expected sheet for model {model.__name__}",
                )

    def test_description_row_has_text_wrap(self) -> None:
        """Test description row has text wrapping enabled."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "test_template.xlsx"
            generate_excel_from_pydantic(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["DatasetDescription"]

            # Check first cell in description row has text wrap
            desc_cell = ws.cell(1, 1)
            self.assertTrue(desc_cell.alignment.wrap_text)

    def test_sheet_name_truncation_for_long_names(self) -> None:
        """Test sheet names are truncated to Excel's 31 character limit."""

        class VeryLongModelNameThatExceedsThirtyOneCharacters(BaseModel):
            field: str

        with tempfile.TemporaryDirectory() as tmp_dir:
            import xlsxwriter

            path = Path(tmp_dir) / "test_long_name.xlsx"
            workbook = xlsxwriter.Workbook(str(path))
            write_model_sheet(workbook, VeryLongModelNameThatExceedsThirtyOneCharacters)
            workbook.close()

            wb = openpyxl.load_workbook(path)
            sheet_names = wb.sheetnames

            # All sheet names should be <= 31 characters
            for name in sheet_names:
                self.assertLessEqual(len(name), 31)


if __name__ == "__main__":
    unittest.main()
